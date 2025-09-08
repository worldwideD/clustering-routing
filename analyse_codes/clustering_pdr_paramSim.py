import multiprocessing
import openpyxl
import os
import argparse
import json
import subprocess
from time import time
import random
from clustering_simulation import getdis, pos

maxwh = 5000
maxd = 2500
stdout_path = 'stdout.txt'

def GetResult(path):
    with open(path, "r") as f:
        str = f.readlines()
    str1 = str[0].split(';')
    str2 = str[1].split(' ')

    result = {}
    for s in str1:
        key = s.split(' = ')[0].strip()
        val = s.split(' = ')[1].strip()
        if key in {"Delivery Ratio", "Pkt Collision Rate", "hello Pkt Collision Rate", "LS Pkt Collision Rate", "Disconnect Drop Ratio"}:
            val = float(val.strip('%'))
        else:
            val = float(val)
        result[key] = val
    result['pdrs'] = str2
    return result

def worker(args):
    # """工作进程函数，执行实际计算任务"""
    # try:
    #     import psutil
    #     p = psutil.Process(os.getpid())
    #     p.cpu_affinity([cpu_id])
    # except ImportError:
    #     print("psutil模块未安装，无法设置CPU亲和性")
    
    # print(f"进程 {os.getpid()} 在CPU {cpu_id} 上处理任务: {task}")
    
    # 执行计算任务
    # result = task ** 2
    command, traceFile = args[0], args[1]
    with open(stdout_path, "w") as f:
        result = subprocess.run(command, stdout=f, stderr=f)
                                
    if result.returncode != 0:
        print(f"❌ 运行失败，错误码: {result.returncode}")
    
    return GetResult(traceFile)

def Simulation(model, long_rate, nodenum, speed, hello, bps, movable_rate, o, g, d, sink, pdrInterval, args):
    configTemplate = f'./src/aqua-sim-ng/examples/settings/{model}/settings_{model}_sink.json'
    # print(f'Reading {configTemplate}.')
    with open(configTemplate, "r") as f:
        data = json.load(f)

    nodenum += sink
    long_rate = int(long_rate)
    simulation_results = []
    trace_dir = f'./tracelog-{model}/{nodenum}_{(int)(float(hello) * nodenum)}_{speed}_{bps}_{long_rate}'
    if model == 'clustering':
        trace_dir = f'./tracelog-{model}-sink/{nodenum}_{(int)(float(hello) * nodenum)}_{speed}_{bps}_{long_rate}'
    if model == 'PAUV':
        trace_dir = f'./tracelog-{model}/{nodenum}_0_0_{bps}_{long_rate}'
    if nodenum <= 30+sink:
        maxwh = 3000
    else:
        maxwh = 5000
    
    argsList = []

    # 分配任务到不同的CPU      
    for runNum in range(args.runs):
        configName = f"./src/aqua-sim-ng/examples/settings/{model}/settings_{model}_{runNum+1}.json"
        outputs = {}
        for key, val in data.items():
            if key == "node_positions":
                positions = pos(nodenum, int((nodenum - sink) * long_rate / 100.) + sink, maxwh, maxd, s=1)
                outputs['node_positions'] = positions
            elif key == 'node_num':
                outputs['node_num'] = nodenum
                
            elif key == 'mobility_models_batch':
                v = val
                v[0]['params']['speed'] = [0.5, speed]
                v[0]['params']['X'] = v[0]['params']['Y'] = [0, maxwh]
                v[0]['params']['Z'] = [0, maxd]
                outputs['mobility_models_batch'] = v
            elif key == 'simulation_time':
                # outputs['simulation_time'] = 100 + 10 * (nodenum - sink - 1) * (nodenum - sink)
                outputs['simulation_time'] = 100 + 10 * 1000
            elif key == 'hello_interval' and model == 'clustering':
                outputs['hello_interval'] = (float)(hello)
            elif key == 'hello_interval':
                outputs['hello_interval'] = (float)(hello)
            elif key == 'omega':
                outputs['omega'] = o#[args.omega, round(1 - 2 * args.omega, 3), args.omega]
            elif key == 'gamma':
                outputs['gamma'] = g#[args.gamma, 1 - args.gamma]
            elif key == 'delta':
                outputs['delta'] = d#[args.delta, 1 - args.delta]
            elif key == 'mu':
                outputs['mu'] = args.mu
            elif key == 'lambda':
                outputs['lambda'] = args.lamb
            # elif key == 'prob' and model == 'LEACH':
            #     outputs['prob'] = prob
            elif key == 'max_speed':
                outputs['max_speed'] = speed
            elif key == 'longNodes_rate':
                outputs['longNodes_rate'] = long_rate / 100.
            elif key == 'forming_time' and model == 'clustering':
                outputs['forming_time'] = 20
            elif key == 'routing_time' and model == 'clustering':
                outputs['routing_time'] = 50
            elif key == 'initialNetworkingCompletion' and model=='clustering':
                outputs['initialNetworkingCompletion'] = 100
            elif key == 'data_rate_bps':
                outputs['data_rate_bps'] = bps
            elif key == 'movable_rate':
                outputs['movable_rate'] = movable_rate
            elif key == 'packet_size':
                outputs['packet_size'] = 50
            else:
                 outputs[key] = val
        if model == 'PAUV':
            outputs['MinX'] = outputs['MinY'] = outputs['MinZ'] = 0
            outputs['MaxX'] = outputs['MaxY'] = maxwh
            outputs['MaxZ'] = maxd
        outputs['sink'] = sink
        outputs['senders'] = [i+1 for i in range(sink, nodenum)]
        outputs['receivers'] = [i+1 for i in range(sink, nodenum)]
        with open(configName, "w") as f:
            json.dump(outputs, f, indent=4, ensure_ascii=False)
                                
        txList = ''
        txset = set()
        txnum = int(nodenum * 0.5)
        for _ in range(txnum):
            src = random.randint(0, nodenum-1)
            dst = random.randint(0, nodenum-1)
            while src == dst or (src * 1000 + dst) in txset:
                src = random.randint(0, nodenum-1)
                dst = random.randint(0, nodenum-1)
            txset.add(src * 1000 + dst)
            txList = txList + str(src) + ',' + str(dst)
            if _ < txnum - 1:
                txList = txList + ';'
        
        senders = ''
        for i in range(nodenum):
            senders = senders + str(i)
        receivers = senders
                                
        command = [
            "./ns3", "run",
            f'dolphin_clustering_sink'
            f' --setting={configName}'
            f' --runNumber={runNum+1}'
            f' --lossrate=0.0'
            f' --packetnum=1000'
            # f' --senders=1'
            # f' --receivers=2'
            f' --printLoggingInfo=False'
            f' --srcDstTxList={txList}'
            f' --regularPDR=1'
            f' --pdrInterval={pdrInterval}'
        ]                         
                                
        traceFile = f'trace_{runNum+1}_1000_0_{nodenum}_{sink+1}_{sink+1}_dolphin_scenario.txt'
        traceFile = os.path.join(trace_dir, traceFile)
        argsList.append((command, traceFile))
    
    # 获取CPU核心数量
    num_cpus = multiprocessing.cpu_count()
    # 创建进程池
    processes = min(int(num_cpus * 0.8), len(argsList))

    # 汇总结果取平均值
    avg_result = {
        'Control Packets Count': 0,
        'Hello Packets Count': 0,
        'LS Packets Count': 0,
        'Switch CH Count': 0,
        'Become CH Count': 0,
        'Become CM Count': 0,
        'Control Overhead(bytes/T)': 0,
        'Delivery Ratio': 0,
        'Data Packets Traffic(bytes)': 0,
        'Avg End2End Delay': 0,
        'Pkt Collision Count': 0,
        'Pkt Collision Rate': 0,
        'hello Pkt Collision Rate': 0,
        'LS Pkt Collision Rate': 0
    }
    pdrs = [0 for _ in range(10)]

    with multiprocessing.Pool(processes=processes) as pool:

        simulation_results = pool.imap_unordered(worker, argsList)
        for result in simulation_results:
            avg_result['Control Packets Count'] += result['Control Packets Count']
            avg_result['Hello Packets Count'] += result['Hello Packets Count']
            avg_result['LS Packets Count'] += result['LS Packets Count']
            avg_result['Switch CH Count'] += result['Switch CH Count']
            avg_result['Become CH Count'] += result['Become CH Count']
            avg_result['Become CM Count'] += result['Become CM Count']
            avg_result['Control Overhead(bytes/T)'] += result['Control Overhead(bytes/T)']
            avg_result['Delivery Ratio'] += result['Delivery Ratio']
            avg_result['Data Packets Traffic(bytes)'] += result['Data Packets Traffic(bytes)']
            avg_result['Avg End2End Delay'] += result['Avg End2End Delay']
            avg_result['Pkt Collision Count'] += result['Pkt Collision Count']
            avg_result['Pkt Collision Rate'] += result['Pkt Collision Rate']
            avg_result['hello Pkt Collision Rate'] += result['hello Pkt Collision Rate']
            avg_result['LS Pkt Collision Rate'] += result['LS Pkt Collision Rate']
            for i in range(10):
                pdrs[i] += float(result['pdrs'][i])

    for k in avg_result:
        avg_result[k] = round(avg_result[k] / args.runs, 3)
    for i in range(10):
        pdrs[i] = round(pdrs[i] / args.runs, 3)
    
    print(avg_result, pdrs)
    return avg_result, pdrs

def main():
    random.seed(time())
    parser = argparse.ArgumentParser()


    # parser.add_argument("--nodenum", default=50, type=int)
    # parser.add_argument("--long_rate", default=50, type=double)
    # parser.add_argument("--max_speed", default=3, type=double)
    # parser.add_argument("--hello_interval", default=60, type=double)
    parser.add_argument("--nodenum", default="50", type=str)
    parser.add_argument("--long_rate", default="50", type=str)
    parser.add_argument("--max_speed", default="3", type=str)
    parser.add_argument("--hello_interval", default="2", type=str)
    parser.add_argument("--bps", default="8", type=str)
    parser.add_argument("--omega", default=0.4, type=float)
    parser.add_argument("--gamma", default=0.5, type=float)
    parser.add_argument("--delta", default=0.5, type=float)
    parser.add_argument("--mu", default=0.2, type=float)
    parser.add_argument("--lamb", default=0.2, type=float)
    parser.add_argument("--runs", default=10, type=int)
    parser.add_argument("--models", default="clustering", type=str)

    args = parser.parse_args()
    models = args.models.split(';')
    nodenums = args.nodenum.split(';')
    longrates = args.long_rate.split(';')
    speeds = args.max_speed.split(';')
    hellos = args.hello_interval.split(';')
    bpss = args.bps.split(';')

    model_default = "clustering"
    # model_default = "LEACH"
    long_rate_default = 50
    nodenum_default = 100
    hello_interval_default = 2
    speed_default = 3
    bps_default = 8

    # for model in models:
    #     Simulation(model, long_rate_default, nodenum_default, speed_default, hello_interval_default, bps_default, args)

    # for long_rate in longrates:
    #     long_rate = int(long_rate)
    #     for nodenum in nodenums:
    #         nodenum = int(nodenum)
    #         Simulation(model_default, long_rate, nodenum, speed_default, hello_interval_default, bps_default, args)

    # for speed in speeds:
    #     speed = int(speed)
    #     for nodenum in nodenums:
    #         nodenum = int(nodenum)
    #         Simulation(model_default, long_rate_default, nodenum, speed, hello_interval_default, bps_default, args)

    # for hello in hellos:
    #     hello = int(hello)
    #     print(f'hello = {hello}')
    #     for nodenum in nodenums:
    #         nodenum = int(nodenum)
    #         Simulation(model_default, long_rate_default, nodenum, speed_default, hello, bps_default, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], args)

    
    # for move in [0, 0.25, 0.5, 0.75, 1]:
    # # for move in [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
    # # for move in [0.5]:
    #     print(f'move ratio = {move}')
    #     for nodenum in nodenums:
    #         nodenum = int(nodenum)
    #         print(f'node num = {nodenum}')
    #         start_t = time()
    #         result = Simulation(args.models, long_rate_default, nodenum, speed_default, args.hello_interval, 8, move, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], 1, args)
    #         vals = [move, nodenum]
    #         for v in result.values():
    #             vals.append(v)
    #         values.append(vals)
    #         end_t = time()
    #         print(f'Total Run Time = {end_t - start_t}(s).')

    for model in ['clustering']:
    # for model in ['LEACH']:
    # for model in [args.models]:
        print(f'model = {model}')
        values = [['maxSpeed',
               'nodenum',
        'Control Packets Count',
        'Hello Packets Count',
        'LS Packets Count',
        'Switch CH Count',
        'Become CH Count',
        'Become CM Count',
        'Control Overhead(bytes/T)',
        'Delivery Ratio',
        'Data Packets Traffic(bytes)',
        'Avg End2End Delay',
        'Pkt Collision Count',
        'Pkt Collision Rate',
        'hello Pkt Collision Rate',
        'LS Pkt Collision Rate']]
        pdrs = []
        # for speed in [1, 2, 3, 4, 5]:
        # # for speed in [5]:
        #     print(f'max speed = {speed}')
        #     for nodenum in nodenums:
        #         nodenum = int(nodenum)
        #         print(f'node num = {nodenum}')
        #         start_t = time()
        #         hello = args.hello_interval
        #         if model != args.models:
        #             hello = 30
        #         if speed >= 1:
        #             result, pdr = Simulation(model, args.long_rate, nodenum, speed, hello, 40, 1.0, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], 1, 1000, args)
        #         else:
        #             result, pdr = Simulation(model, args.long_rate, nodenum, 100, hello, 40, 0, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], 1, 1000, args)  
        #         vals = [speed, nodenum]
        #         for v in result.values():
        #             vals.append(v)
        #         values.append(vals)
        #         pdrs.append(pdr)
        #         end_t = time()
        #         print(f'Total Run Time = {end_t - start_t}(s).')
        for hello in [1, 1.5, 2, 2.5, 3]:
        # for speed in [5]:
            print(f'hello = {hello}')
            speed = 5
            for nodenum in nodenums:
                nodenum = int(nodenum)
                print(f'node num = {nodenum}')
                start_t = time()
                if model != args.models:
                    hello = 30
                if speed >= 1:
                    result, pdr = Simulation(model, args.long_rate, nodenum, speed, hello, 40, 1.0, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], 1, 1000, args)
                else:
                    result, pdr = Simulation(model, args.long_rate, nodenum, 100, hello, 40, 0, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], 1, 1000, args)  
                vals = [speed, nodenum]
                for v in result.values():
                    vals.append(v)
                values.append(vals)
                pdrs.append(pdr)
                end_t = time()
                print(f'Total Run Time = {end_t - start_t}(s).')

        index = len(values)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(0, index):
            for j in range(0, len(values[i])):
                sheet.cell(row=i+1, column=j+1, value=values[i][j])
        sheet = workbook.create_sheet('PDRs')
        index = len(pdrs)
        for i in range(0, index):
            for j in range(0, 10):
                sheet.cell(row=i+1, column=j+1, value=pdrs[i][j])
        workbook.save(f'result_pdr_{model}.xlsx')

    # for nodenum in nodenums:
    #     nodenum = int(nodenum)
    #     Simulation(model_default, long_rate_default, nodenum, speed_default, hello_interval_default, bps_default, [0.4, 0.2, 0.4], [0.5, 0.5], [0.5, 0.5], args)
    
    # for p in [0.3, 0.4, 0.5]:
    #     Simulation(model_default, long_rate_default, nodenum_default, speed_default, hello_interval_default, bps_default, p, args)

    # for bps in bpss:
    #     bps = int(bps)
    #     for nodenum in nodenums:
    #         nodenum = int(nodenum)
    #         Simulation(model_default, long_rate_default, nodenum, speed_default, hello_interval_default, bps, args)
                            
    

if __name__ == "__main__":
    main()