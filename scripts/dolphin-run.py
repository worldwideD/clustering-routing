import argparse
import subprocess
import re
from openpyxl import Workbook

BroadNum = []
LeaderRecvACK = []
LeaderRecvAllACK = []
FirstBroadToNodesNum = []
FirstBroadToNodesDeliveryRate = []
FirstBroadToNodesTimeDR100 = []
FirstBroadToNodesTimeDR90 = []
FirstBroadToNodesTimeDR80 = []

def run_simulation(simulation_id,leader_id):
    """
    Run the NS-3 simulation with the given run ID.
    Each simulation's output is redirected to a file named after the run number.
    """
    # Define the command to run the NS-3 simulation with appropriate parameters
    ns3_command = [
        "./ns3", "run", 
        f'dolphin_scenario_1 --setting=/home/hull/settings_scenario_1.json --runNumber={simulation_id}'
    ]

    # Open a file to redirect the output of the simulation
    with open(f"simulation_output_run_{simulation_id}.txt", "w") as output_file:
        # Run the NS-3 simulation and redirect both stdout and stderr to the output file
        subprocess.run(ns3_command, stdout=output_file, stderr=output_file)

    print(f"Run {simulation_id} completed. Output saved to simulation_output_run_{simulation_id}.txt")
    
    input_file = f"trace_{simulation_id}.txt"
    command = ['python3', 'src/aqua-sim-ng/scripts/dolphin-apptrace-2.py', f'--filepath={input_file}', f'--leaderid={leader_id}']
    print(f"Run apptrace-2.py {input_file} ");
    result = subprocess.run(command, capture_output=True, text=True)
    print(result.stdout)
    
    #BroadNum
    numbers = re.findall(r'BN:\s*(\d+)', result.stdout)
    if numbers:
        BroadNum.append(int(numbers[0]))
    else:
        BroadNum.append(0)

    #LeaderRecvACK
    numbers = re.findall(r'ACKN:\s*(\d+)', result.stdout)
    if numbers:
        LeaderRecvACK.append(int(numbers[0]))
    else:
        LeaderRecvACK.append(0)

    #LeaderRecvAllACK
    numbers = re.findall(r'ALLACK:\s*(\d+)', result.stdout)
    if numbers:
        LeaderRecvAllACK.append(int(numbers[0]))
    else:
        LeaderRecvAllACK.append(0)

    #FirstBroadToNodesNum
    numbers = re.findall(r'FirstBTN:\s*(\d+)', result.stdout)
    if numbers:
        FirstBroadToNodesNum.append(int(numbers[0]))
    else:
        FirstBroadToNodesNum.append(0)

    #FirstBroadToNodesDeliveryRate
    numbers = re.findall(r'BDR:\s*(\d+\.\d{1,2})', result.stdout)
    if numbers:
        FirstBroadToNodesDeliveryRate.append(float(numbers[0]))
    else:
        FirstBroadToNodesDeliveryRate.append(0)

    #FirstBroadToNodesTimeDR80
    numbers = re.findall(r'BC80:\s*(\d+\.\d{1,2})', result.stdout)
    if numbers:
        FirstBroadToNodesTimeDR80.append(float(numbers[0]))
    else:
        FirstBroadToNodesTimeDR80.append(0)

    #FirstBroadToNodesTimeDR90
    numbers = re.findall(r'BC90:\s*(\d+\.\d{1,2})', result.stdout)
    if numbers:
        FirstBroadToNodesTimeDR90.append(float(numbers[0]))
    else:
        FirstBroadToNodesTimeDR90.append(0)

    #FirstBroadToNodesTimeDR100
    numbers = re.findall(r'BC:\s*(\d+\.\d{1,2})', result.stdout)
    if numbers:
        FirstBroadToNodesTimeDR100.append(float(numbers[0]))
    else:
        FirstBroadToNodesTimeDR100.append(0)

def writeExcel():
    wb = Workbook()

    # 激活默认工作表
    ws = wb.active

    # 写入标题
    ws.append(["BroadNum", "DeliveryRate"])

    for num, id_ in zip(BroadNum, FirstBroadToNodesDeliveryRate):
       ws.append([num, id_])

    # 保存 Excel 文件
    wb.save("output.xlsx")

def runByLeadid(num_runs,leader_id,ws):
    print(f"Starting {num_runs} simulation runs...,leader_id:{leader_id}")

    # Run the simulation for the specified number of times
    for i in range(1, num_runs + 1):
        run_simulation(i,leader_id)

    row_index = leader_id+1
    column_index = 1

    #写excel
    # 写入标题
    ws.append(["Leader","BroadNum", "LeaderRecvAllACK","DeliveryRate","DeliveryRate80%","DeliveryRate90%","DeliveryRate100%"])
    ws.cell(row=row_index, column=column_index, value=leader_id)
    column_index+=1

    #输出结果
    print("*********result********")
    filtered_times = [time for time in BroadNum if time != 0]
    # 计算非零元素的平均值
    if filtered_times:
        average_time = sum(filtered_times) / len(filtered_times)
    else:
        average_time = 0
    max_value = max(BroadNum) if BroadNum else None
    min_value = min(BroadNum) if BroadNum else None
    print(f"广播次数: {BroadNum}, Average: {average_time:.2f}, Max: {max_value}, Min: {min_value}")
    ws.cell(row=row_index, column=column_index, value=f"{average_time:.2f}")
    column_index+=1

    print("领航节点收到ACK数量：", LeaderRecvACK)
    print("领航节点是否收到所有ACK：", LeaderRecvAllACK)
    filtered_times = [time for time in LeaderRecvAllACK]
    # 计算非零元素的平均值
    if filtered_times:
        average_time = sum(filtered_times) / len(filtered_times)
    else:
        average_time = 0
    ws.cell(row=row_index, column=column_index, value=average_time)
    column_index+=1
    print("所有其它节点收到第一次广播的次数：", FirstBroadToNodesNum)

    average_bdr = sum(FirstBroadToNodesDeliveryRate)/len(FirstBroadToNodesDeliveryRate)
    max_value = max(FirstBroadToNodesDeliveryRate) if FirstBroadToNodesDeliveryRate else None
    min_value = min(FirstBroadToNodesDeliveryRate) if FirstBroadToNodesDeliveryRate else None
    print(f"全网广播送达率: {FirstBroadToNodesDeliveryRate}, Average: {average_bdr}%, Max: {max_value}%, Min: {min_value}%");
    ws.cell(row=row_index, column=column_index, value=average_bdr)
    column_index+=1

    # 过滤掉值为 0 的元素
    filtered_times = [time for time in FirstBroadToNodesTimeDR80 if time != 0]
    # 计算非零元素的平均值
    if filtered_times:
        average_time = sum(filtered_times) / len(filtered_times)
    else:
        average_time = 0
    max_value = max(FirstBroadToNodesTimeDR80) if FirstBroadToNodesTimeDR80 else None
    min_value = min(FirstBroadToNodesTimeDR80) if FirstBroadToNodesTimeDR80 else None
    print(f"送达率80%下，全网广播完成时间(s): {FirstBroadToNodesTimeDR80}, Average: {average_time:.2f}, Max: {max_value}, Min: {min_value}")
    ws.cell(row=row_index, column=column_index, value=f"{average_time:.2f}")
    column_index+=1

    # 过滤掉值为 0 的元素
    filtered_times = [time for time in FirstBroadToNodesTimeDR90 if time != 0]
    # 计算非零元素的平均值
    if filtered_times:
        average_time = sum(filtered_times) / len(filtered_times)
    else:
        average_time = 0
    max_value = max(FirstBroadToNodesTimeDR90) if FirstBroadToNodesTimeDR90 else None
    min_value = min(FirstBroadToNodesTimeDR90) if FirstBroadToNodesTimeDR90 else None
    print(f"送达率90%下，全网广播完成时间(s): {FirstBroadToNodesTimeDR90}, Average: {average_time:.2f}, Max: {max_value}, Min: {min_value}")
    ws.cell(row=row_index, column=column_index, value=f"{average_time:.2f}")
    column_index+=1

    # 过滤掉值为 0 的元素
    filtered_times = [time for time in FirstBroadToNodesTimeDR100 if time != 0]
    # 计算非零元素的平均值
    if filtered_times:
        average_time = sum(filtered_times) / len(filtered_times)
    else:
        average_time = 0
    max_value = max(FirstBroadToNodesTimeDR100) if FirstBroadToNodesTimeDR100 else None
    min_value = min(FirstBroadToNodesTimeDR100) if FirstBroadToNodesTimeDR100 else None
    print(f"送达率100%下，全网广播完成时间(s): {FirstBroadToNodesTimeDR100}, Average: {average_time:.2f}, Max: {max_value}, Min: {min_value}")
    ws.cell(row=row_index, column=column_index, value=f"{average_time:.2f}")
    column_index+=1

def parse_range(value):
    try:
        # 检查输入格式是否为 'start-end'
        start, end = map(int, value.split('-'))
        if start > end:
            raise argparse.ArgumentTypeError(f"无效范围: {value}")
        return range(start, end + 1)
    except ValueError:
        raise argparse.ArgumentTypeError(f"无效范围: {value}")

def main():
    # Create an argument parser to handle command-line input
    parser = argparse.ArgumentParser(description="Run NS-3 simulations multiple times.")

    # Add an argument to specify the number of runs
    parser.add_argument("--runs", type=int, default=1, help="Number of simulation runs (default: 1)")
    parser.add_argument("--leaderid", type=int, default=3, help="leader id")
    parser.add_argument("--leaders", type=parse_range, help="leader range: 1-10")

    # Parse the command-line arguments
    args = parser.parse_args()

    # Get the number of simulation runs
    num_runs = args.runs
    leader_id = args.leaderid
    leaders = []

    #init excel
    wb = Workbook()
    # 激活默认工作表
    ws = wb.active

    if args.leaders is not None:
        leaders = list(args.leaders)
        print(f"leader range: {leaders}")
        for leader in leaders:
            runByLeadid(num_runs,leader,ws)
    else:
        runByLeadid(num_runs,leader_id,ws)

    wb.save(f"output-{leader_id}.xlsx")

if __name__ == "__main__":
    main()